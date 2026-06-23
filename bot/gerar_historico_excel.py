def gerar_historico_excel():

    import sqlite3
    import os
    from datetime import datetime
    from openpyxl import Workbook

    DB_PATH = "relatorios/historico.db"

    UNIDADES_ABAS = {
        "1.1.1": "SINOP-MT",
        "1.1.2": "DOURADOS-MS",
        "1.1.3": "NOVA MUTUM-MT",
        "1.1.6": "SIDROLANDIA-MS",
        "1.1.8": "BALSAS-MA",
        "1.1.10": "LUIS E. MAGALHAES-BA",
    }

    HEADERS = [
        "ID","Unidade","CanceladoEm","TipoNoShow","Motivo",
        "Tipo Frete","Agendado","Cliente","Periodo",
        "Codigo","Descricao","Cpf","Nome","Placas","DataExecucao",
    ]

    os.makedirs("relatorios", exist_ok=True)

    mes = datetime.now().strftime("%Y-%m")
    excel_path = f"relatorios/historico_noshow_{mes}.xlsx"

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM noshow")
    dados = cursor.fetchall()

    conn.close()

    wb = Workbook()

    default = wb.active
    wb.remove(default)

    resumo = wb.create_sheet("RESUMO")
    resumo.append(["Unidade", "Total"])

    abas = {}

    for codigo, nome in UNIDADES_ABAS.items():
        ws = wb.create_sheet(nome)
        ws.append(HEADERS)
        abas[codigo] = ws

    for row in dados:
        unidade_codigo = row[1]
        ws = abas.get(unidade_codigo)
        if ws:
            ws.append(row)

    for codigo, ws in abas.items():
        total = ws.max_row - 1
        resumo.append([UNIDADES_ABAS[codigo], total])

    wb.save(excel_path)

    print("Histórico Excel gerado:", excel_path)

    return excel_path