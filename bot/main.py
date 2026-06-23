from playwright.sync_api import sync_playwright

from .utils.excel_formatter import formatar_planilha
from .gerar_historico_excel import gerar_historico_excel
from .config import load_config
from .context import BotContext
from .report.realtime import criar_relatorios_execucao
from .ui.apex import safe_networkidle, fechar_alert_sucesso
from .stages import NoShowStage, Produto41Stage, CancelamentoStage
from .utils.email_smtp import enviar_email_smtp
from .report.database import criar_tabela

from openpyxl import load_workbook
from datetime import datetime
import os


PIPELINE = [
    CancelamentoStage(),
    NoShowStage(),
]


def realizar_login(ctx: BotContext):
    page = ctx.page
    cfg = ctx.config

    page.on("dialog", lambda d: d.accept())

    page.goto(cfg.url, wait_until="networkidle")
    page.fill("#P9999_USERNAME", cfg.usuario)
    page.fill("#P9999_PASSWORD", cfg.senha)
    page.click("#B142316728014002480")

    safe_networkidle(page)
    page.wait_for_timeout(1200)
    fechar_alert_sucesso(page)

    ctx.log("✅ Login realizado com sucesso")


def main():

    cfg = load_config()

    # ============================================================
    # LOG FILE (NOVO)
    # ============================================================

    os.makedirs("logs", exist_ok=True)
    log_path = os.path.join(
        "logs",
        f"execucao_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.log"
    )

    # cria banco SQLite se não existir
    criar_tabela()

    # ============================================================
    # RELATÓRIO DA EXECUÇÃO
    # ============================================================

    xlsx_path, jsonl_path, wb, ws = criar_relatorios_execucao()

    print("📄 Relatório XLSX:", xlsx_path)
    print("🧾 Backup JSONL:", jsonl_path)

    # ============================================================

    with sync_playwright() as p:

        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        page.set_default_timeout(cfg.default_timeout_ms)
        page.set_default_navigation_timeout(cfg.default_timeout_ms)

    
        ctx = BotContext(
            page=page,
            wb=wb,
            ws=ws,
            xlsx_path=xlsx_path,
            jsonl_path=jsonl_path,
            config=cfg,
            log_path=log_path,
        )

        ctx.log("🚀 Iniciando execução do bot")

        try:

            realizar_login(ctx)

            for stage in PIPELINE:

                try:
                    fechar_alert_sucesso(page)
                except Exception:
                    pass

                try:
                    safe_networkidle(page)
                except Exception:
                    pass

                ctx.log(f"\n▶️ Iniciando stage: {stage.name}")

                try:
                    res = stage.run(ctx)

                except Exception as e:
                    ctx.log(f"❌ Stage {stage.name} lançou exceção: {repr(e)}")
                    break

                ctx.log(
                    f"✅ Stage {res.name}: ok={res.ok} total={res.total} "
                    f"{('- ' + res.details) if res.details else ''}"
                )

                if not res.ok:
                    ctx.log(f"⛔ Parando pipeline: stage {stage.name} falhou.")
                    break

        finally:

            # ============================================================
            # SALVAR RELATÓRIO DIÁRIO
            # ============================================================

            try:
                formatar_planilha(ws)
                wb.save(xlsx_path)
                ctx.log("📄 Relatório diário salvo")
            except Exception as e:
                ctx.log(f"❌ Erro ao salvar XLSX: {repr(e)}")

            # ============================================================
            # GERAR HISTÓRICO
            # ============================================================

            try:
                historico_path = gerar_historico_excel()
                ctx.log(f"📊 Histórico gerado: {historico_path}")

                if historico_path and os.path.exists(historico_path):
                    wb_hist = load_workbook(historico_path)

                    for ws_hist in wb_hist.worksheets:
                        formatar_planilha(ws_hist)

                    wb_hist.save(historico_path)
                    ctx.log("📊 Histórico formatado com sucesso")

            except Exception as e:
                ctx.log(f"❌ Erro ao gerar histórico: {repr(e)}")
                historico_path = None

            # ============================================================
            # VALIDAÇÃO DE DADOS
            # ============================================================

            tem_noshow = ws.max_row > 1
            tem_cancelamento = False
            cancelamento_path = None

            if hasattr(ctx, "state") and "cancelamentos_xlsx_path" in ctx.state:
                cancelamento_path = ctx.state["cancelamentos_xlsx_path"]

                if cancelamento_path and os.path.exists(cancelamento_path):
                    try:
                        wb_cancel = load_workbook(cancelamento_path)
                        ws_cancel = wb_cancel.active

                        if ws_cancel.max_row > 1:
                            tem_cancelamento = True

                        formatar_planilha(ws_cancel)
                        wb_cancel.save(cancelamento_path)

                        ctx.log("📄 Planilha de cancelamentos formatada")

                    except Exception as e:
                        ctx.log(f"⚠️ Erro ao validar cancelamentos: {repr(e)}")

            # ============================================================
            # ENVIO DE EMAIL
            # ============================================================

            try:

                arquivos = []

                if tem_cancelamento:
                    arquivos.append(cancelamento_path)

                if tem_noshow:
                    arquivos.append(xlsx_path)

                if historico_path:
                    arquivos.append(historico_path)

                enviar_email_smtp(
                    arquivos,
                    tem_noshow=tem_noshow,
                    tem_cancelamento=tem_cancelamento
                )

                ctx.log("📧 Email enviado com sucesso")

            except Exception as e:
                ctx.log(f"❌ Erro ao enviar email: {repr(e)}")

            # ============================================================
            # FECHAR BROWSER
            # ============================================================

            try:
                browser.close()
                ctx.log("🔒 Browser fechado")
            except Exception:
                pass

    ctx.log("\n💾 Finalizado.")
    ctx.log(f"   XLSX: {xlsx_path}")
    ctx.log(f"   JSONL: {jsonl_path}")
    ctx.log(f"   HISTÓRICO: {historico_path}")


if __name__ == "__main__":
    main()
