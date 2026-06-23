import os
import json
from .database import inserir_noshow
from datetime import datetime
from openpyxl import Workbook, load_workbook


# adicionamos DataExecucao
REPORT_HEADERS = [
    "DataExecucao",
    "Unidade",
    "CanceladoEm",
    "TipoNoShow",
    "Motivo",
    "ID",
    "Tipo Frete",
    "Agendado",
    "Cliente",
    "Periodo",
    "Codigo",
    "Descricao",
    "Cpf",
    "Nome",
    "Placas",
]


UNIDADES_ABAS = {
    "1.1.1": "SINOP-MT",
    "1.1.2": "DOURADOS-MS",
    "1.1.3": "NOVA MUTUM-MT",
    "1.1.6": "SIDROLANDIA-MS",
    "1.1.8": "BALSAS-MA",
    "1.1.10": "LUIS E. MAGALHAES-BA",
}


# ============================================================
# RELATÓRIO DA EXECUÇÃO
# ============================================================

def criar_relatorios_execucao():

    os.makedirs("relatorios", exist_ok=True)

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    xlsx_path = os.path.join("relatorios", f"cancelamento_noshow_{stamp}.xlsx")
    jsonl_path = os.path.join("relatorios", f"cancelamento_noshow_{stamp}.jsonl")

    wb = Workbook()

    ws = wb.active
    ws.title = "NoShow"

    ws.append(REPORT_HEADERS)

    ws._last_unidade = None

    wb.save(xlsx_path)

    with open(jsonl_path, "a", encoding="utf-8") as f:
        f.write("")

    return xlsx_path, jsonl_path, wb, ws


def anexar_linha(ws, wb, xlsx_path, row):

    ws.append(row)
    wb.save(xlsx_path)


def anexar_jsonl(path, registro):

    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(registro, ensure_ascii=False) + "\n")
        f.flush()


# ============================================================
# HISTÓRICO
# ============================================================

def obter_historico_noshow():

    mes = datetime.now().strftime("%Y-%m")

    path = os.path.join("relatorios", f"historico_noshow_{mes}.xlsx")

    if os.path.exists(path):

        wb = load_workbook(path)

    else:

        wb = Workbook()

        default = wb.active
        wb.remove(default)

        resumo = wb.create_sheet("RESUMO")
        resumo.append(["Unidade", "Total"])

        for nome in UNIDADES_ABAS.values():

            ws = wb.create_sheet(nome)

            ws.append(REPORT_HEADERS)

        wb.save(path)

    return path, wb


# ============================================================
# DUPLICAÇÃO
# ============================================================

def registro_ja_existe(ws, registro):

    id_registro = str(registro.get("ID"))

    for row in ws.iter_rows(min_row=2, values_only=True):

        if str(row[5]) == id_registro:
            return True

    return False


# ============================================================
# HISTÓRICO
# ============================================================

def anexar_historico_noshow(path, wb, registro):

    unidade = (registro.get("Unidade") or "").upper()

    aba_nome = None

    for chave, aba in UNIDADES_ABAS.items():

        if chave in unidade:
            aba_nome = aba
            break

    if not aba_nome:
        return

    ws = wb[aba_nome]

    # evitar duplicação
    if registro_ja_existe(ws, registro):
        return

    registro["DataExecucao"] = datetime.now().strftime("%Y-%m-%d")

    ws.append([registro.get(h, "") for h in REPORT_HEADERS])

    wb.save(path)

    atualizar_resumo(wb, path)


# ============================================================
# RESUMO
# ============================================================

def atualizar_resumo(wb, path):

    resumo = wb["RESUMO"]

    resumo.delete_rows(2, resumo.max_row)

    for aba in UNIDADES_ABAS.values():

        ws = wb[aba]

        total = ws.max_row - 1

        resumo.append([aba, total])

    wb.save(path)


# ============================================================
# GRAVAÇÃO
# ============================================================

def anexar_cancelamento_e_persistir(
    wb,
    ws,
    xlsx_path,
    jsonl_path,
    registro,
    historico_wb=None,
    historico_path=None
):

    from datetime import datetime

    registro["DataExecucao"] = datetime.now().strftime("%Y-%m-%d")

    anexar_linha(
        ws,
        wb,
        xlsx_path,
        [registro.get(h, "") for h in REPORT_HEADERS],
    )

    anexar_jsonl(jsonl_path, registro)

    # salvar no banco
    if registro.get("TipoNoShow"):

        inserir_noshow(registro)

        # if historico_wb:
        #     anexar_historico_noshow(
        #         historico_path,
        #         historico_wb,
        #         registro
        #     )