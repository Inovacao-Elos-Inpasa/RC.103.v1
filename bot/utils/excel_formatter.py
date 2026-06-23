from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def formatar_planilha(ws):
    # ============================================================
    # 🎨 ESTILO DO HEADER
    # ============================================================

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ============================================================
    # 🔄 FILTRO AUTOMÁTICO
    # ============================================================

    ws.auto_filter.ref = ws.dimensions

    # ============================================================
    # 🧊 CONGELAR HEADER
    # ============================================================

    ws.freeze_panes = "A2"

    # ============================================================
    # 🎯 ZEBRA (linhas alternadas)
    # ============================================================

    fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        fill = fill_1 if row % 2 == 0 else fill_2

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # ============================================================
    # 📏 AJUSTE AUTOMÁTICO DE COLUNAS
    # ============================================================

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2